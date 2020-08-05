#! python3
# Usage: python spreadsheet_to_text_files.py <spreadsheet.xlsx>

import openpyxl
from pathlib import Path
import sys
import os

# If a spreadsheet is specified, store in variable
if len(sys.argv) > 1:
    spreadsheet = sys.argv[1]

# Create a Path object as a working directory
currentDir = Path(os.path.abspath(os.curdir))
inputFile = currentDir / spreadsheet
print('Reading contents from ' + str(inputFile) + '...')

wb = openpyxl.load_workbook(inputFile)
sheet = wb.active

# Iterate through columns first, creating a new textfile for each column.
for column in range(1, sheet.max_column+1):
    with open(f'output{column}.txt', 'w') as f:
        for row in range(1, sheet.max_row+1):
            # If the cell is not blank, write the value to a new line
            if sheet.cell(row=row, column=column).value:
                f.write(sheet.cell(row=row, column=column).value)

