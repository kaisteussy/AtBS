import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active

lst = []

for row in range(1, sheet.max_row+1):
    lst.append([])
    for column in range(1, sheet.max_column+1):
        lst[row-1].append(sheet.cell(row=row, column=column).value)

updated_wb = openpyxl.Workbook()
newSheet = updated_wb.active

print(lst)

for y, column in enumerate(lst):
    for x, row in enumerate(column):
        newSheet.cell(row=x+1, column=y+1).value = lst[y][x]

updated_wb.save('updated-produceSales.xlsx')
wb.close()
updated_wb.close()
