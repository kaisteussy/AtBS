import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active

wb_output = openpyxl.Workbook()
output_sheet = wb_output.active

for x in range(1, sheet.max_row + 1):
    for y in range(1, sheet.max_column + 1):
        output_sheet.cell(row=y, column=x).value = sheet.cell(row=x, column=y).value

wb_output.save('updated-produceSales.xlsx')
wb_output.close()
wb.close()
