#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for each country
import pprint
import census2010
import os

import openpyxl

print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb[wb.sheetnames[0]]
countyData = {}

print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})

    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)

print('Writing results...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')



