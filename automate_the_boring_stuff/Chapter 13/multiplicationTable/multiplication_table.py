#! python3
# Usage: python multiplication_table.py <number>

import openpyxl
import sys

# Get the number from the command line argument to create a multiplication table for
if len(sys.argv) == 2:
    number = int(sys.argv[1])

print(len(sys.argv))

# Create the workbook and store the active sheet in a variable
wb = openpyxl.Workbook()
sheet = wb.active

# Fill rows starting at 1, ending at the number input
for row in range(1, number+1):
    sheet['A' + str(row+1)] = row

# Fill columns starting at 1, ending at the number input
for column in range(1, number+1):
    sheet.cell(row=1, column=column+1).value = column

# Fill in the multiplication table
for row in range(1, number+1):
    for column in range(1, number+1):
        sheet.cell(row=row+1,column=column+1).value = row*column

# Save the workbook
wb.save('multiplicationtable.xlsx')